import sys

import jinja2 as jj
import openpyxl
import yaml


def twitter(value):
    name = value
    icon = "ti-twitter-alt"
    link = f"https://twitter.com/{value}"
    return name, icon, link


def phone(value):
    name = value
    icon = "ti-mobile"
    link = f"tel:{value}"
    return name, icon, link


def flickr(value):
    name = value
    icon = "ti-flickr"
    link = f"https://flickr.com/photos/{value}"
    return name, icon, link


def youtube(value):
    name = "YouTube Channel"
    icon = "ti-youtube"
    link = value
    return name, icon, link


def facebook(value):
    name = value
    icon = "ti-facebook"
    link = f"https://facebook.com/{value}"
    return name, icon, link


def email(value):
    name = f"{value}@okayama-u.ac.jp"
    icon = "ti-email"
    link = f"mailto:{value}@okayama-u.ac.jp"
    return name, icon, link


def ResearcherID(value):
    name = f"ResearcherID {value}"
    icon = "ti-id-badge"
    link = f"https://researcherid.com/rid/{value}"
    return name, icon, link


def ORCID(value):
    name = f"ORCID {value}"
    icon = "ti-id-badge"
    link = f"https://orcid.org/{value}"
    return name, icon, link


def Scholar(value):
    name = f"Scholar {value}"
    icon = "ti-google"
    link = f"https://scholar.google.co.jp/citations?user={value}"
    return name, icon, link


ranks = {
    "教授": "Professor",
    "准教授": "Assoc. Professor",
    "助教": "Assist. Professor",
    "講師": "Lecturer",
}
groups = {
    "分子化学": "Molecular Chemistry",
    "物質化学": "Material Chemistry",
    "反応化学": "Reaction Chemistry",
    "界面化学": "Interfacial Chemistry",
}
labs = {
    "ナノ化学": ["nano", "ナノ化学研究室", "Nano Chemistry Laboratory"],
    "錯体化学研究室": ["coord", "錯体化学研究室", "Coordination Chemistry Laboratory"],
    "機能有機化学研究室": [
        "funcchem",
        "機能有機化学研究室",
        "Functional Organic Chemistry Laboratory",
    ],
    "無機化学研究室": ["inorganic", "無機化学研究室", "Inorganic Chemistry Laboratory"],
    "有機化学研究室": ["organic", "有機化学研究室", "Laboratory of Organic Chemistry"],
    "表面物理化学研究室": ["surfphys", "表面物理化学研究室", "Surface Physical Chemistry"],
    "理論化学研究室": ["theochem", "理論化学研究室", "Theoretical Chemistry Laboratory"],
    "理論物理化学研究室": [
        "theophyschem",
        "理論物理化学研究室",
        "Theoretical Physical Chemistry Laboratory",
    ],
    "理論計算化学研究室": [
        "theocomp",
        "理論計算化学研究室",
        "Theoretical and Computational Chemistry Laboratory",
    ],
    "界面化学研究室": [
        "interface",
        "界面化学研究室",
        "Interfacial Chemistry Laboratory",
    ],
    "反応有機化学研究室": [
        "synth",
        "反応有機化学研究室",
        "Synthetic and Physical Organic Chemistry Laboratory",
    ],
    "分析化学研究室": [
        "analytical",
        "分析化学研究室",
        "Analytical Chemistry Laboratory",
    ],
    "粉体化学研究室": [
        "fine",
        "粉体化学研究室",
        "Fine Powder and Surface Chemistry Laboratory",
    ],
    "有機合成化学研究室": [
        "synorg",
        "有機合成化学研究室",
        "Synthetic Organic Chemistry Laboratory",
    ],
    "分光化学研究室": [
        "spectro",
        "分光化学研究室",
        "Spectrochemistry Laboratory",
    ],
}

uid = sys.argv[1]
wb = openpyxl.load_workbook(f"{uid}.xlsx")
ws = wb.worksheets[0]

keys = [x[0].value for x in ws["A1:A22"]]
values = [x[0].value for x in ws["B1:B22"]]
d = dict(zip(keys, values))
print(d)

keys = [x[0].value for x in ws["A23:A34"]]
values = [x[0].value for x in ws["B23:B34"]]
dc = dict(zip(keys, values))
dc["email"] = d["e-mail"]
dc["phone"] = d["研究室電話"]


da = dict()
da["ResearcherID"] = d["ResearcherID"]
da["ORCID"] = d["ORCID"]
da["Scholar"] = d["Google Scholar "]

contacts = []
for key, value in dc.items():
    if value is not None:
        name, icon, link = eval(f"{key}('{value}')")
        contacts.append(dict(name=name, icon=icon, link=link))

d["contact_yaml"] = yaml.dump({"contact": contacts})
d["image"] = f"images/faculty/{uid}.jpg"

achieve = []
for key, value in da.items():
    if value is not None:
        name, icon, link = eval(f"{key}('{value}')")
        achieve.append(dict(name=name, icon=icon, link=link))

d["achievements_yaml"] = yaml.dump({"achievements": achieve})

del d[None]
d["description"] = d["1行説明"]
d["description_en"] = d["one liner(en)"]
d["name"] = d["name(en)"]
d["tag1_en"] = d["tag1(en)"]
d["tag2_en"] = d["tag2(en)"]
d["tag3_en"] = d["tag3(en)"]
d["rank"] = ranks[d["職階"]]
d["group"] = groups[d["領域"]]
labnames = labs[d["研究室"]]
d["研究室_en"] = labnames[2]
d["lab_ja"] = yaml.dump(
    {"laboratory": {"id": labnames[0], "name": labnames[1]}},
    allow_unicode=True,
)
d["lab_en"] = yaml.dump(
    {"laboratory": {"id": labnames[0], "name": labnames[2]}}
)

with open("faculty.ja.md") as f:
    jjt = f.read()
template = jj.Template(jjt)
s = template.render(d)
with open(f"{sys.argv[1]}.ja.md", "w") as f:
    f.write(s)


with open("faculty.en.md") as f:
    jjt = f.read()
template = jj.Template(jjt)
s = template.render(d)
with open(f"{sys.argv[1]}.en.md", "w") as f:
    f.write(s)

print(s)
