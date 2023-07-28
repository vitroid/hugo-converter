import sys

import jinja2 as jj
import openpyxl
import yaml


def twitter(value):
    name = value
    icon = "ti-twitter-alt"
    link = f"https://twitter.com/{value}"
    return name, icon, link


def facebook(value):
    name = value
    icon = "ti-facebook"
    link = f"https://facebook.com/{value}"
    return name, icon, link


def email(value):
    name = f"{value}@okayama-u.ac.jp"
    icon = "ti-email"
    link = f"mailto:{value}@okayama-u.ac.jp"
    return name, icon, link


def ResearcherID(value):
    name = f"ResearcherID {value}"
    icon = "ti-id-badge"
    link = f"https://researcherid.com/rid/{value}"
    return name, icon, link


def ORCID(value):
    name = f"ORCID {value}"
    icon = "ti-id-badge"
    link = f"https://orcid.org/{value}"
    return name, icon, link


def Scholar(value):
    name = f"Scholar {value}"
    icon = "ti-google"
    link = f"https://facebook.com/{value}"
    return name, icon, link


def phone(value):
    name = value
    icon = "ti-mobile"
    link = f"tel:{value}"
    return name, icon, link


def fax(value):
    name = f"FAX {value}"
    icon = "ti-printer"
    link = f"tel:{value}"
    return name, icon, link


wb = openpyxl.load_workbook(f"{sys.argv[1]}.xlsx")
ws = wb.worksheets[0]

keys = [x[0].value for x in ws["A1:A16"]]
values = [x[0].value for x in ws["B1:B16"]]
d = dict(zip(keys, values))
print(d)

keys = [x[0].value for x in ws["A17:A26"]]
values = [x[0].value for x in ws["B17:B26"]]
dc = dict(zip(keys, values))
dc["email"] = d["代表e-mail"]
dc["phone"] = d["代表電話"]
dc["fax"] = d["Fax"]


# list of faculties
keys = [x[0].value for x in ws["A27:A29"]]
values_ja = [x[0].value for x in ws["B27:B29"]]
values_en = [x[0].value for x in ws["C27:C29"]]
f_ja = [
    {"id": key, "name": name}
    for key, name in zip(keys, values_ja)
    if key is not None
]
f_en = [
    {"id": key, "name": name}
    for key, name in zip(keys, values_en)
    if key is not None
]

contacts = []
for key, value in dc.items():
    if value is not None:
        name, icon, link = eval(f"{key}('{value}')")
        contacts.append(dict(name=name, icon=icon, link=link))

d["contact_yaml"] = yaml.dump({"contact": contacts})

del d[None]
d["description"] = d["1行説明"]
d["description_en"] = d["one liner(en)"]
d["name_en"] = d["研究室(en)"]
d["tag1_en"] = d["tag1(en)"]
d["tag2_en"] = d["tag2(en)"]
d["tag3_en"] = d["tag3(en)"]
groups = {
    "分子化学": "Molecular Chemistry",
    "物質化学": "Material Chemistry",
    "反応化学": "Reaction Chemistry",
    "界面化学": "Interfacial Chemistry",
}
d["group"] = groups[d["領域"]]


with open("laboratory.ja.md") as f:
    jjt = f.read()
template = jj.Template(jjt)
d["faculties"] = yaml.dump({"faculties": f_ja}, allow_unicode=True)
s = template.render(d)
with open(f"{sys.argv[1]}.ja.md", "w") as f:
    f.write(s)

with open("laboratory.en.md") as f:
    jjt = f.read()
template = jj.Template(jjt)
s = template.render(d)
d["faculties"] = yaml.dump({"faculties": f_en}, allow_unicode=True)
s = template.render(d)
with open(f"{sys.argv[1]}.en.md", "w") as f:
    f.write(s)

print(s)
